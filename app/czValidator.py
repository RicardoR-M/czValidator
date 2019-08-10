import io
from time import time

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from app.conf import sheet_name
from app.validaciones import usuario, texto, fecha, numero, no_vacio, monitor_cz, sino, fcr, no_si, detalle_nofcr, vacio, obs_recomendacion, sn


def validador(tipo, value):
    if tipo == 'sn':
        if sn(value):
            print('Error: ' + tipo)
    elif tipo == 'usuario':
        if usuario(value):
            print('Error: ' + tipo)
    elif tipo == 'texto':
        if texto(value):
            print('Error: ' + tipo)
    elif tipo == 'fecha':
        if fecha(value):
            print('Error: ' + tipo)
    elif tipo == 'numero':
        if numero(value):
            print('Error: ' + tipo)
    elif tipo == 'no_vacio':
        if no_vacio(value):
            print('Error: ' + tipo)
    elif tipo == 'monitor_cz':
        if monitor_cz(value):
            print('Error: ' + tipo)
    elif tipo == 'sino':
        if sino(value):
            print('Error: ' + tipo)
    elif tipo == 'fcr':
        if fcr(value):
            print('Error: ' + tipo)
    elif tipo == 'no_si':
        if no_si(value):
            print('Error: ' + tipo)
    elif tipo == 'detalle_nofcr':
        if detalle_nofcr(value):
            print('Error: ' + tipo)
    elif tipo == 'vacio':
        if vacio(value):
            print('Error: ' + tipo)
    elif tipo == 'obs_recomendacion':
        if obs_recomendacion(value):
            print('Error:' + tipo)
    else:
        raise RuntimeError('Validacion no implementada: ' + tipo)


def check_cz(excel):
    validaciones = [dict(cabecera='CODIGO DEL ASESOR', col='A', tipo='usuario'),
                    dict(cabecera='NOMBRE DEL ASESOR', col='B', tipo='texto'),
                    dict(cabecera='FECHA DE MONITOREO', col='F', tipo='fecha'),
                    dict(cabecera='SN - IPCC', col='H', tipo='sn'),
                    dict(cabecera='FECHA DE LA LLAMADA', col='I', tipo='fecha'),
                    dict(cabecera='NOMBRE DEL CLIENTE', col='L', tipo='texto'),
                    dict(cabecera='NRO° TELFONO DEL QUE LLAMA (MSISDN )', col='M', tipo='numero'),
                    dict(cabecera=' NRO°  INDENTIFICADO  (TELEFONO EN CONSULTA)', col='M', tipo='numero'),
                    dict(cabecera='TIPO DE LLAMADA', col='O', tipo='texto'),
                    dict(cabecera='MOTIVO DE CONSULTA', col='P', tipo='no_vacio'),
                    dict(cabecera='NOMBRE DEL MONITOR', col='Z', tipo='monitor_cz'),
                    dict(cabecera='SUB  CALIFICACIÓN', col='BJ', tipo='numero'),
                    dict(cabecera='CALIFICACIÓN FINAL', col='BP', tipo='numero'),
                    dict(cabecera='FCR', col='BV', tipo='sino'),
                    dict(cabecera='RESPONSABILIDAD DE NO FCR', col='BW', tipo='fcr'),
                    dict(cabecera='MOTIVO DE NO FCR', col='BX', tipo='no_si'),
                    dict(cabecera='PROCESO CLARO NO FCR', col='BY', tipo='no_si'),
                    dict(cabecera='DETALLE DE NO FCR', col='BZ', tipo='detalle_nofcr'),
                    dict(cabecera='ACCION QUE REALIZO  EL ASESOR', col='CA', tipo='no_si'),
                    dict(cabecera='NPS', col='CB', tipo='vacio'),
                    dict(cabecera='TNPS ', col='CC', tipo='vacio'),
                    dict(cabecera='OBSERVACION / RECOMENDACIÓN', col='CE', tipo='obs_recomendacion'),
                    dict(cabecera='Semana de llamada', col='CF', tipo='vacio'),
                    dict(cabecera='Semana de monitoreo', col='CF', tipo='vacio'),
                    dict(cabecera='Adherencia al proceso', col='CH', tipo='sino')
                    ]
    tini = time()

    wb = load_workbook(filename=excel, data_only=True, keep_links=False, keep_vba=False, read_only=True)

    if sheet_name() not in wb.sheetnames:
        raise RuntimeError('No se encontro la hoja Data')

    ws = wb[sheet_name()]

    lista_errores = []

    for i, row in enumerate(ws.rows):
        if i == 0:  # Salta la cabecera
            continue
        # if i == 2:
        #     for item in validaciones:
        #         print(f'{item["cabecera"]} - {row[column_index_from_string(item["col"]) - 1].value} - {type(row[column_index_from_string(item["col"]) - 1].value)}')
        for item in validaciones:
            validador(item['tipo'], row[column_index_from_string(item['col']) - 1].value)

    # for item in validaciones:
    #     print(f'{item["col"]} - {column_index_from_string(item["col"])}')

    tend = time()
    print(f'Timpo: {tend - tini}')


if __name__ == '__main__':
    with open('tec.xlsx', 'rb') as f:
        mem_file = io.BytesIO(f.read())

    check_cz(mem_file)
