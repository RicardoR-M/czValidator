import io
from time import time

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from app.conf import sheet_name, col_fecha_monitoreo, col_fecha_llamada
from app.plantillas import ErrorValidacion
from app.validaciones import usuario, texto, fecha, numero, no_vacio, monitor_cz, sino, fcr, no_si, detalle_nofcr, vacio, obs_recomendacion, sn, fechas_cz


def validador(item_dict, value, fila):
    if item_dict['tipo'] == 'sn':
        if sn(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='SN no contiene solo números')
    elif item_dict['tipo'] == 'usuario':
        if usuario(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Usuario invalido')
    elif item_dict['tipo'] == 'texto':
        if texto(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda contiene números o caracteres inválidos')
    elif item_dict['tipo'] == 'fecha':
        if fecha(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda no contiene una fecha válida')
    elif item_dict['tipo'] == 'numero':
        if numero(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda contiene letras o caracteres inválidos')
    elif item_dict['tipo'] == 'no_vacio':
        if no_vacio(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda no se encuentra vacia')
    elif item_dict['tipo'] == 'monitor_cz':
        if monitor_cz(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Nombre de monitor Cruzado no coincide con el asignado en el programa')
    elif item_dict['tipo'] == 'sino':
        if sino(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda con valores inválidos')
    elif item_dict['tipo'] == 'fcr':
        if fcr(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda no contiene un responsable válido de FCR')
    elif item_dict['tipo'] == 'no_si':
        if no_si(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda contiene un valor inválido')
    elif item_dict['tipo'] == 'detalle_nofcr':
        if detalle_nofcr(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda Detalle de No FCR contiene el caracter ">"')
    elif item_dict['tipo'] == 'vacio':
        if vacio(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda contiene valores cuando debe estar vacia')
    elif item_dict['tipo'] == 'obs_recomendacion':
        if obs_recomendacion(value):
            return ErrorValidacion(fila=fila, col=item_dict['col'], cabecera=item_dict['cabecera'], valor=value, msg='Celda Observaciones contiene guiones ---')
    else:
        raise RuntimeError('Validacion no implementada: ' + item_dict['tipo'])

    return False


def check_cz(excel):
    validaciones = [dict(cabecera='CODIGO DEL ASESOR', col='A', tipo='usuario'),
                    dict(cabecera='NOMBRE DEL ASESOR', col='B', tipo='texto'),
                    dict(cabecera='FECHA DE MONITOREO', col='F', tipo='fecha'),
                    dict(cabecera='SN - IPCC', col='H', tipo='sn'),
                    dict(cabecera='FECHA DE LA LLAMADA', col='I', tipo='fecha'),
                    dict(cabecera='NOMBRE DEL CLIENTE', col='L', tipo='texto'),
                    dict(cabecera='NRO° TELFONO DEL QUE LLAMA (MSISDN )', col='M', tipo='numero'),
                    dict(cabecera=' NRO°  INDENTIFICADO  (TELEFONO EN CONSULTA)', col='N', tipo='numero'),
                    dict(cabecera='TIPO DE LLAMADA', col='O', tipo='texto'),
                    dict(cabecera='MOTIVO DE CONSULTA', col='P', tipo='no_vacio'),
                    dict(cabecera='NOMBRE DEL MONITOR', col='Z', tipo='monitor_cz'),
                    dict(cabecera='SUB  CALIFICACIÓN', col='AZ', tipo='numero'),  # Varia en cuanto a tec
                    dict(cabecera='CALIFICACIÓN FINAL', col='BF', tipo='numero'),  # varia en cuanto a tec
                    dict(cabecera='FCR', col='BL', tipo='sino'),  # varia en cuanto a tec
                    dict(cabecera='RESPONSABILIDAD DE NO FCR', col='BM', tipo='fcr'),  # Varia
                    dict(cabecera='MOTIVO DE NO FCR', col='BN', tipo='no_si'),  # varia
                    dict(cabecera='PROCESO CLARO NO FCR', col='BO', tipo='no_si'),  # Varia
                    dict(cabecera='DETALLE DE NO FCR', col='BP', tipo='detalle_nofcr'),  # Varia
                    dict(cabecera='ACCION QUE REALIZO  EL ASESOR', col='BQ', tipo='no_si'),  # Varia
                    dict(cabecera='NPS', col='BR', tipo='vacio'),
                    dict(cabecera='TNPS ', col='BS', tipo='vacio'),
                    dict(cabecera='OBSERVACION / RECOMENDACIÓN', col='BU', tipo='obs_recomendacion'),
                    # dict(cabecera='Semana de llamada', col='CH', tipo='vacio'),
                    # dict(cabecera='Semana de monitoreo', col='CI', tipo='vacio'),
                    # dict(cabecera='Adherencia al proceso', col='CJ', tipo='sino')
                    ]
    tini = time()

    wb = load_workbook(filename=excel, data_only=True, keep_links=False, keep_vba=False, read_only=True)

    if sheet_name() not in wb.sheetnames:
        raise RuntimeError('No se encontro la hoja Data')

    ws = wb[sheet_name()]

    lista_errores = []
    for i, row in enumerate(ws.rows):
        if i == 0:  # Salta la cabecera
            continue

        for item in validaciones:
            # Validaciones generales de tipo y contenido de celdas
            validacion = validador(item, row[column_index_from_string(item['col']) - 1].value, i + 1)
            # Si encuentra un error lo adjunta a la lista
            if validacion:
                lista_errores.append(validacion)

        # valida si las fechas de monitoreo y llamada excede los dos días, si son iguales o si la fecha de monitoreo es mayor
        moni_vs_llamada = fechas_cz(monitoreo=row[column_index_from_string(col_fecha_monitoreo()) - 1].value, llamada=row[column_index_from_string(col_fecha_llamada()) - 1].value, fila=i + 1)
        if moni_vs_llamada:
            lista_errores.append(moni_vs_llamada)

    if len(lista_errores) > 0:
        for error in lista_errores:
            print(f'Fila: {error.fila} Col: {error.col} // Cabecera: {error.cabecera} // Val: "{error.valor}" // Msg: {error.msg}')

    tend = time()

    print(f'Timpo: {tend - tini}')

    return


if __name__ == '__main__':
    with open('post.xlsx', 'rb') as f:
        mem_file = io.BytesIO(f.read())

    check_cz(mem_file)
