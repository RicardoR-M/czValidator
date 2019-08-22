import io

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from conf import sheet_name, col_sn, col_fecha_monitoreo, col_fecha_llamada
from plantillas import ErrorValidacion


def comparator_cz(informe_anterior, informe_now):
    wb_anterior = load_workbook(filename=informe_anterior, data_only=True, keep_links=False, keep_vba=False, read_only=True)
    wb_now = load_workbook(filename=informe_now, data_only=True, keep_links=False, keep_vba=False, read_only=True)

    if sheet_name() not in wb_anterior.sheetnames or sheet_name() not in wb_now.sheetnames:
        raise RuntimeError('No se encontro la hoja Data')

    ws_anterior = wb_anterior[sheet_name()]
    ws_now = wb_now[sheet_name()]

    dict_anterior = {}
    dict_now = {}

    # carga la informacion de los ws en memoria
    for i, row in enumerate(ws_anterior):
        if i == 0:  # Omite la cabecera
            continue
        dict_anterior[row[column_index_from_string(col_sn()) - 1].value] = dict(f_mon=row[column_index_from_string(col_fecha_monitoreo()) - 1].value,
                                                                                f_llamada=row[column_index_from_string(col_fecha_llamada()) - 1].value,
                                                                                fila=i + 1)
    for i, row in enumerate(ws_now):
        if i == 0:  # Omite la cabecera
            continue
        dict_now[row[column_index_from_string(col_sn()) - 1].value] = dict(f_mon=row[column_index_from_string(col_fecha_monitoreo()) - 1].value,
                                                                           f_llamada=row[column_index_from_string(col_fecha_llamada()) - 1].value,
                                                                           fila=i + 1)

    # valida si la cantidad de evaluaciones del reporte anterior es menor que el actual
    if len(dict_anterior) > len(dict_now):
        raise RuntimeError('Error: El informe anterior contiene más evaluaciones que el nuevo reporte')

    lista_errores = []
    for sn in dict_anterior:
        if sn in dict_now:
            if dict_anterior[sn]['f_mon'] != dict_now[sn]['f_mon']:
                lista_errores.append(ErrorValidacion(fila=f'Anterior: {dict_anterior[sn]["fila"]} - Actual: {dict_now[sn]["fila"]}',
                                                     col=sn,
                                                     cabecera='F.Monitoreo anterior vs actual',
                                                     valor=f'Ant: {dict_anterior[sn]["f_mon"]} - Now: {dict_now[sn]["f_mon"]}',
                                                     msg='Se encontro diferente fecha de monitoreo en los informes'))
            if dict_anterior[sn]['f_llamada'] != dict_now[sn]['f_llamada']:
                lista_errores.append(ErrorValidacion(fila=f'Anterior: {dict_anterior[sn]["fila"]} - Actual: {dict_now[sn]["fila"]}',
                                                     col=sn,
                                                     cabecera='F.Llamada anterior vs actual',
                                                     valor=f'Ant: {dict_anterior[sn]["f_llamada"]} - Now: {dict_now[sn]["f_llamada"]}',
                                                     msg='Se encontro diferente fecha de llamada en los informes'))
    if len(lista_errores) > 0:
        for error in lista_errores:
            print(f'Filas: ({error.fila}) // SN: {error.col} // Cabecera: {error.cabecera} // Valores: ({error.valor}) // Msg: {error.msg}')
    else:
        print('No se encontro ninguna observación')


if __name__ == '__main__':
    with open('post_anterior.xlsx', 'rb') as f:
        mem_file_anterior = io.BytesIO(f.read())

    with open('post_now.xlsx', 'rb') as f:
        mem_file_now = io.BytesIO(f.read())

    comparator_cz(mem_file_anterior, mem_file_now)
